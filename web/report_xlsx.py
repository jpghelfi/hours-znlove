"""The per-project Excel export: one sheet per project, people then their log.

Kept out of app.py (which holds routes/HTTP concerns) and out of notion_ops.py
(which holds Notion reads/writes): this module only turns entries that have
already been read into a workbook.

Shape, per project sheet: a title block, a "by person" table (hours, days,
entries), then every entry with its date, person and comment — the thing you
hand to a client or a PM, rather than the flat row dump /project.csv gives.
"""
from __future__ import annotations

import io
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

_INK = "1B2138"
_MUTED = "6B7280"
_HEAD_FILL = PatternFill("solid", fgColor="EEF2F9")
_TITLE_FONT = Font(bold=True, size=15, color=_INK)
_SUB_FONT = Font(size=10, color=_MUTED)
_SECTION_FONT = Font(bold=True, size=11, color=_INK)
_TH_FONT = Font(bold=True, size=10, color=_INK)
_TOTAL_FONT = Font(bold=True, size=10, color=_INK)
_RULE = Border(bottom=Side(style="thin", color="D6DBE6"))
_HOURS_FMT = "0.##"

# Excel forbids these in a sheet name, and caps it at 31 chars
_BAD_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")


def sheet_title(name: str, used: set) -> str:
    """A legal, unique sheet name for a project. Long project names collide once
    truncated to 31 chars, so uniqueness is enforced with a numeric suffix."""
    base = _BAD_SHEET_CHARS.sub(" ", name).strip() or "Project"
    title = base[:31]
    n = 2
    while title.lower() in used:
        suffix = f" ({n})"
        title = base[:31 - len(suffix)] + suffix
        n += 1
    used.add(title.lower())
    return title


def _widths(ws, widths: dict) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def group(entries: list[dict]) -> list[dict]:
    """entries -> one record per project, each with its people and its log.

    Grouped by project *id* (two projects may share a name), ordered by hours
    desc so the busiest project is the first sheet; people inside a project are
    ordered the same way.
    """
    groups: dict = {}
    for e in entries:
        g = groups.setdefault(e.get("project_id") or e.get("project") or "(none)", {
            "name": e.get("project") or "(none)", "hours": 0.0, "people": {},
            "goals": {}, "log": [],
        })
        g["hours"] += e["hours"]
        g["log"].append(e)
        p = g["people"].setdefault(e.get("person_id") or e["person"], {
            "name": e["person"], "hours": 0.0, "entries": 0, "days": set(),
        })
        p["hours"] += e["hours"]
        p["entries"] += 1
        p["days"].add(e["date"])
        # goals are grouped by name here: within one project a name is one
        # goal, and the sheet is read by people who know the names, not the ids
        gl = g["goals"].setdefault(e.get("goal") or "",
                                   {"name": e.get("goal") or "Unassigned",
                                    "hours": 0.0, "entries": 0})
        gl["hours"] += e["hours"]
        gl["entries"] += 1
    out = []
    for g in sorted(groups.values(), key=lambda g: (-g["hours"], g["name"].lower())):
        out.append({
            "name": g["name"],
            "hours": round(g["hours"], 2),
            "entries": len(g["log"]),
            "days": len({e["date"] for e in g["log"]}),
            # only worth a table when the project actually uses goals: an
            # all-Unassigned block would be a row saying "no goals" twice
            "goals": sorted(
                ({"name": gl["name"], "hours": round(gl["hours"], 2),
                  "entries": gl["entries"],
                  "unassigned": gl["name"] == "Unassigned"}
                 for gl in g["goals"].values()),
                key=lambda gl: (gl["unassigned"], -gl["hours"], gl["name"].lower()))
            if any(k for k in g["goals"]) else [],
            "people": sorted(
                ({"name": p["name"], "hours": round(p["hours"], 2),
                  "entries": p["entries"], "days": len(p["days"])}
                 for p in g["people"].values()),
                key=lambda p: (-p["hours"], p["name"].lower())),
            # chronological: the sheet reads as a work log, oldest first
            "log": sorted(g["log"], key=lambda e: (e["date"], e["person"].lower())),
        })
    return out


def _summary_sheet(wb: Workbook, groups: list[dict], period_label: str) -> None:
    ws = wb.create_sheet("Summary")
    ws["A1"] = "Hours by project"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = period_label
    ws["A2"].font = _SUB_FONT
    headers = ["Project", "Hours", "People", "Days", "Entries"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=i, value=h)
        c.font, c.fill, c.border = _TH_FONT, _HEAD_FILL, _RULE
    row = 5
    for g in groups:
        ws.cell(row=row, column=1, value=g["name"])
        ws.cell(row=row, column=2, value=g["hours"]).number_format = _HOURS_FMT
        ws.cell(row=row, column=3, value=len(g["people"]))
        ws.cell(row=row, column=4, value=g["days"])
        ws.cell(row=row, column=5, value=g["entries"])
        row += 1
    ws.cell(row=row, column=1, value="Total").font = _TOTAL_FONT
    total = ws.cell(row=row, column=2, value=round(sum(g["hours"] for g in groups), 2))
    total.font, total.number_format = _TOTAL_FONT, _HOURS_FMT
    ws.cell(row=row, column=5, value=sum(g["entries"] for g in groups)).font = _TOTAL_FONT
    _widths(ws, {"A": 42, "B": 10, "C": 10, "D": 8, "E": 10})
    ws.freeze_panes = "A5"


def _project_sheet(wb: Workbook, g: dict, period_label: str, used: set) -> None:
    ws = wb.create_sheet(sheet_title(g["name"], used))
    ws["A1"] = g["name"]
    ws["A1"].font = _TITLE_FONT
    n = len(g["people"])
    ws["A2"] = (f"{period_label} · {g['hours']:g} h · {n} {'person' if n == 1 else 'people'} · "
                f"{g['entries']} {'entry' if g['entries'] == 1 else 'entries'}")
    ws["A2"].font = _SUB_FONT

    ws["A4"] = "By person"
    ws["A4"].font = _SECTION_FONT
    for i, h in enumerate(["Person", "Hours", "Days", "Entries"], start=1):
        c = ws.cell(row=5, column=i, value=h)
        c.font, c.fill, c.border = _TH_FONT, _HEAD_FILL, _RULE
    row = 6
    for p in g["people"]:
        ws.cell(row=row, column=1, value=p["name"])
        ws.cell(row=row, column=2, value=p["hours"]).number_format = _HOURS_FMT
        ws.cell(row=row, column=3, value=p["days"])
        ws.cell(row=row, column=4, value=p["entries"])
        row += 1
    ws.cell(row=row, column=1, value="Total").font = _TOTAL_FONT
    tot = ws.cell(row=row, column=2, value=g["hours"])
    tot.font, tot.number_format = _TOTAL_FONT, _HOURS_FMT

    if g["goals"]:
        # what the month went into — usually the line a client reads first, so
        # it sits above the log; only on a sheet whose project uses goals
        row += 2
        ws.cell(row=row, column=1, value="By goal").font = _SECTION_FONT
        row += 1
        for i, h in enumerate(["Goal", "Hours", "Entries"], start=1):
            c = ws.cell(row=row, column=i, value=h)
            c.font, c.fill, c.border = _TH_FONT, _HEAD_FILL, _RULE
        row += 1
        for gl in g["goals"]:
            ws.cell(row=row, column=1, value=gl["name"])
            ws.cell(row=row, column=2, value=gl["hours"]).number_format = _HOURS_FMT
            ws.cell(row=row, column=3, value=gl["entries"])
            row += 1
        ws.cell(row=row, column=1, value="Total").font = _TOTAL_FONT
        gt = ws.cell(row=row, column=2, value=g["hours"])
        gt.font, gt.number_format = _TOTAL_FONT, _HOURS_FMT

    log_head = row + 3
    ws.cell(row=log_head - 1, column=1, value="Log").font = _SECTION_FONT
    # the Goal column exists only when the project uses goals, so an export
    # from a project that doesn't is exactly the file it was before
    heads = ["Date", "Person", "Hours"] + (["Goal"] if g["goals"] else []) + ["Comment", "Ticket"]
    for i, h in enumerate(heads, start=1):
        c = ws.cell(row=log_head, column=i, value=h)
        c.font, c.fill, c.border = _TH_FONT, _HEAD_FILL, _RULE
    desc_col = 5 if g["goals"] else 4
    row = log_head + 1
    for e in g["log"]:
        ws.cell(row=row, column=1, value=e["date"])
        ws.cell(row=row, column=2, value=e["person"])
        ws.cell(row=row, column=3, value=e["hours"]).number_format = _HOURS_FMT
        if g["goals"]:
            ws.cell(row=row, column=4, value=e.get("goal") or "")
        c = ws.cell(row=row, column=desc_col, value=e["description"])
        # comments are free text and often long: wrap rather than run off the page
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if e.get("task_url"):
            # the linked Notion ticket, clickable — the column stays empty for
            # every entry logged without one
            t = ws.cell(row=row, column=desc_col + 1, value=e.get("task") or "Notion ticket")
            t.hyperlink, t.style = e["task_url"], "Hyperlink"
        row += 1
    last = chr(ord("A") + desc_col)      # the Ticket column: E, or F with goals
    if g["log"]:
        ws.auto_filter.ref = f"A{log_head}:{last}{row - 1}"
    _widths(ws, {"A": 12, "B": 24, "C": 9, "D": 26, "E": 90, "F": 34} if g["goals"]
                else {"A": 12, "B": 24, "C": 9, "D": 90, "E": 34})
    # keep the log's header visible while scrolling a long month
    ws.freeze_panes = f"A{log_head + 1}"


def build(entries: list[dict], period_label: str, scope_label: str) -> bytes:
    """The workbook as bytes: a Summary sheet (when more than one project) plus
    one sheet per project. An empty period still yields a valid one-sheet file —
    Excel refuses to open a workbook with no sheets."""
    groups = group(entries)
    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet; every sheet here is named
    if not groups:
        ws = wb.create_sheet("No entries")
        ws["A1"] = f"No hours logged · {scope_label}"
        ws["A1"].font = _TITLE_FONT
        ws["A2"] = period_label
        ws["A2"].font = _SUB_FONT
        _widths(ws, {"A": 60})
    else:
        if len(groups) > 1:
            _summary_sheet(wb, groups, period_label)
        used: set = set()
        for g in groups:
            _project_sheet(wb, g, period_label, used)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
